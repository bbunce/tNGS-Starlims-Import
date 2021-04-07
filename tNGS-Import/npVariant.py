import openpyxl
import re

class Variant:

    def __init__(self, ws_main, row):
        self.ws_main = ws_main
        self.row = row
        # self.gender = gender - need to figure out how to get this variable

        if self.ws_main.cell(row=self.row, column=4).value != None and \
                self.ws_main.cell(row=self.row, column=5).value != None:
            self.variantPresent = True
            self.codingEffect = self.ws_main.cell(row=self.row, column=9).value
            self.variantType = self.ws_main.cell(row=self.row, column=10).value
            self.variantLoc = self.ws_main.cell(row=self.row, column=11).value
            self.variantStatus = self.get_variantStatus()
            self.chr = self.ws_main.cell(row=self.row, column=23).value
            self.gene = self.get_gene()
            self.genotype = self.get_genotype()
            self.gNom = self.get_gNom()
            self.transcript = self.ws_main.cell(row=self.row, column=45).value
            self.exon = self.ws_main.cell(row=self.row, column=12).value
            self.intron = self.ws_main.cell(row=self.row, column=43).value
            self.cNom = self.get_cNom()
            self.pNom = self.get_pNom()
        elif self.ws_main.cell(row=self.row, column=4).value != None and \
                self.ws_main.cell(row=self.row, column=5).value == None:
            self.gene = self.get_gene()
            self.variantPresent = True
        else:
            self.variantPresent = False

    def get_gene(self):
        gene = self.ws_main.cell(row=self.row, column=4).value
        if gene.find('_') == -1:
            return gene
        else:
            return gene[:gene.find('_')]

    def get_genotype(self):
        try:
            genotype = float(self.ws_main.cell(row=self.row, column=5).value)
        except ValueError:
            genotype = str(self.ws_main.cell(row=self.row, column=5).value)
        except TypeError:
            return None

        # todo look in detail re: mitochondrial gene formats
        # if self.chr != 'X' or self.chr != 'Y' or not self.gene.startswith('mt'):
        if self.chr != 'X' or self.chr != 'Y':
            if isinstance(genotype, str):
                if genotype == "0/1":
                    return "Heterozygous"
                if genotype == "1/1":
                    return "Homozygous"
            else:
                if (genotype >= 0.45 and genotype <= 0.55):
                    return "Heterozygous"
                if (genotype >= 0.9 and genotype <= 1.1):
                    return "Homozygous"
            # elif (self.chr == 'X' or self.chr == 'Y') and self.gender == 'male':
            #     if genotype == "1/1" or (genotype >= 0.9 and genotype <= 1.1):
            #         return "Hemizygous"
            # elif self.gene.startswith('mt'):
            #     return "Heteroplasmic level?"
        else:
            return "Check Gender/Hemizygous"

    def get_gNom(self):
        gNom = {"gStart": "", "gEnd": "", "gFull": ""}
        gNom["gStart"] = self.ws_main.cell(row=self.row, column=43).value
        gNom["gEnd"] = self.ws_main.cell(row=self.row, column=44).value
        gNom["gFull"] = re.split(":g.", self.ws_main.cell(row=self.row, column=6).value)[1]
        return gNom

    def get_cNom(self):
        cNom = {"cStart": "", "cEnd": "", "cFull": "", "cRef": "", "cAlt": "", "cIndel": ""}
        cNom["cFull"] = re.split(":c.", self.ws_main.cell(row=self.row, column=7).value)[1]
        if cNom["cFull"].find("_") != -1:
            cNom["cStart"] = re.findall("[0-9]+", cNom["cFull"])[0]
            cNom["cEnd"] = re.findall("[0-9]+", cNom["cFull"])[1]
        else:
            cNom["cStart"] = re.findall("[0-9]+", cNom["cFull"])[0]
            cNom["cEnd"] = re.findall("[0-9]+", cNom["cFull"])[0]

        if cNom["cFull"].find(">") != -1:
            cNom["cRef"] = re.split("[>]", cNom["cFull"])[0][-1]
            cNom["cAlt"] = re.split("[>]", cNom["cFull"])[1]
        if cNom["cFull"].find("_") != -1:
            cNom["cIndel"] = re.split("[a-z]+", cNom["cFull"])[-1]
        return cNom

    def get_pNom(self):
        if self.variantLoc == "intron":
            return "p.?"

        aminoAcid = self.ws_main.cell(row=self.row, column=8).value
        pNom = {"pStart": "", "pEnd": "", "pFull": "", "pRef": "", "pAlt": "", "pIndel": ""}
        pNom["pFull"] = aminoAcid[2:].replace("(", "").replace(")", "")
        if pNom["pFull"].find("_") != -1:
            pNom["pStart"] = re.findall("[0-9]+", aminoAcid)[0]
            pNom["pEnd"] = re.findall("[0-9]+", aminoAcid)[1]
        else:
            pNom["pStart"] = re.findall("[0-9]+", aminoAcid)[0]
            pNom["pEnd"] = re.findall("[0-9]+", aminoAcid)[0]

        pNom["pRef"] = re.findall("[a-zA-Z]{3}", aminoAcid)[0]
        if pNom["pFull"].find("=") != -1:
            pNom["pAlt"] = pNom["pRef"]
        elif pNom["pFull"].find("*") != -1:
            pNom["pAlt"] = "Ter"
        else:
            pNom["pAlt"] = re.findall("[a-zA-Z]{3}", aminoAcid)[1]

        # todo pNom Indel not working
        # if pNom["pFull"].find("_") != -1:
        #     for x in re.findall("[a-zA-Z]{3}", aminoAcid)[2:]:
        #         print(type(x), x)
        #         if x != "del" or x != "ins":
        #             pNom["pIndel"] += x
        return pNom

    def get_variantStatus(self):
        status = self.ws_main.cell(row=self.row, column=2).value
        if status == "Novel variant detected" or status == "Confirmation reqd" or status.startswith("Sanger"):
            return "Uncertain"
        elif status.startswith("Variant detected"):
            return "Likely pathogenic/Pathogenic"
        else:
            return status
