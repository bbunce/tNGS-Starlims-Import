import openpyxl
from datetime import datetime
import csv, re

class ImportWorksheet:

    def __init__(self, varPath, varID, starID, workDir):
        self.varPath = varPath
        self.varID = varID
        self.starID = starID
        self.workDir = workDir
        self.timeNow = datetime.now().strftime("%Y-%m-%d %H%M")
        self.newWbName = f"{self.workDir}{self.varID}_{self.starID} tNGS-StarImp {self.timeNow}.xlsx"
        self.wb = self.createCopy()
        self.ws_main = self.wb.get_sheet_by_name("STARLiMS_import")
        self.ws_main_noRows = self.ws_main.max_row
        self.createWorksheets()

    def createCopy(self):
        wb = openpyxl.load_workbook(self.varPath)
        sheets = wb.sheetnames
        for s in sheets:
            if s != "STARLiMS_import":
                sheet_name = wb.get_sheet_by_name(s)
                wb.remove_sheet(sheet_name)
        wb.save(self.newWbName)
        return openpyxl.load_workbook(self.newWbName)

    def createWorksheets(self):
        # Create new sheet and populate header for mutation surveyor import file
        ws_ms = self.wb.create_sheet("MutationSurveyor")
        header_ms = ["Sample Name", "Reference Name", "Lane Quality", "ROI Coverage", "#nts below threshold",
                    "Quality ROI", "Variant1", "Variant2", "Variant3", "Variant4"]
        ws_ms.cell(row=1, column=1).value = "Warning! Mutation calls outside of GBK/SEQ or conflicting GBK/SEQ," \
                                            + " report may not be displayed properly!"
        for i in range(len(header_ms)):
            ws_ms.cell(row=2, column=i+1).value = header_ms[i]

        # Create new sheet and populate header for variant details import file
        ws_var = self.wb.create_sheet("VariantDetails")
        header_var = ["Well", "Sample", "VariantDetails",
                      "Gene1", "Zygosity1", "Inheritance1", "HGVS1", "Genomic1", "Pathogenicity1",
                      "Gene2", "Zygosity2", "Inheritance2", "HGVS2", "Genomic2", "Pathogenicity2",
                      "Gene3", "Zygosity3", "Inheritance3", "HGVS3", "Genomic3", "Pathogenicity3"]
        for i in range(len(header_var)):
            ws_var.cell(row=1, column=i+1).value = header_var[i]
        self.wb.save(self.newWbName)

    def aa_conv(self, aa3):
        aa_dict = {'Ala': 'A', 'Arg': 'R', 'Asn': 'N', 'Asp': 'D', 'Cys': 'C',
                   'Glu': 'E', 'Gln': 'Q', 'Gly': 'G', 'His': 'H', 'Ile': 'I', 'Leu': 'L',
                   'Lys': 'K', 'Met': 'M', 'Phe': 'F', 'Pro': 'P', 'Ser': 'S', 'Thr': 'T',
                   'Trp': 'W', 'Tyr': 'Y', 'Val': 'V', 'Ter': 'X', '?': '?'}
        return aa_dict[aa3]

    def write_variantDetails(self, sample_var):
        ws_var = self.wb.get_sheet_by_name("VariantDetails")
        for i, sample in enumerate(sample_var):
            ws_var.cell(row=i+2, column=1).value = sample_var[sample][0]
            ws_var.cell(row=i+2, column=2).value = sample

            variant = ""
            col_count = 4
            for x in range(2, len(sample_var[sample])):
                sampleInfo = sample_var[sample][x]
                if sampleInfo.variantPresent is True and sampleInfo.confirmationRqd is False:
                    try:
                        if sampleInfo.genotype[:3] == "Hem":
                            genotype = "Hem/Hom"
                        elif sampleInfo.genotype is None:
                            genotype = "?/?"
                        else:
                            genotype = sampleInfo.genotype[:3]
                    except:
                        genotype = "?/?"
                    try:
                        variant += f"{sampleInfo.gene} {genotype} " \
                                   f"p.{sampleInfo.pNom} c.{sampleInfo.cNom['cFull']}; "
                        ws_var.cell(row=i + 2, column=col_count).value = sampleInfo.gene
                        ws_var.cell(row=i + 2, column=col_count + 1).value = sampleInfo.genotype
                        ws_var.cell(row=i + 2, column=col_count + 2).value = "Maternal/Paternal/Biparental/De novo"
                        ws_var.cell(row=i + 2, column=col_count + 3).value = \
                            f"{sampleInfo.transcript}:c.{sampleInfo.cNom['cFull']} p.{sampleInfo.pNom}"
                        ws_var.cell(row=i + 2, column=col_count + 4).value = \
                            f"Chr{sampleInfo.chr}(GRCh37):g.{sampleInfo.gNom['gFull']}"
                        ws_var.cell(row=i + 2, column=col_count + 5).value = sampleInfo.variantStatus
                        col_count += 6
                    except:
                        print(sample)
                elif sampleInfo.variantPresent is True and sampleInfo.confirmationRqd is True:
                    variant += f"{sampleInfo.gene} confirmation/in-fill; "

            ws_var.cell(row=i+2, column=3).value = variant
        self.wb.save(self.newWbName)
        self.to_csv("VariantDetails", "csv", ",")


    def write_mutationSurveyor(self, sample_var):
        ws_ms = self.wb.get_sheet_by_name("MutationSurveyor")
        for i, sample in enumerate(sample_var):
            instr_id = sample_var[sample][1]
            instr_id_sub = "_" + re.findall("_[A-Za-z _]+[v0-9]*_", instr_id)[0][1:-1].replace("_", " ") + "_"
            instr_id = re.sub("_[A-Za-z _]+[v0-9]*_", instr_id_sub, instr_id)
            ws_ms.cell(row=3+i, column=1).value = instr_id

            col_count = 7
            for x in range(2, len(sample_var[sample])):
                sampleInfo = sample_var[sample][x]
                if sampleInfo.variantPresent is True and sampleInfo.confirmationRqd is False:
                    ws_ms.cell(row=i + 3, column=col_count).value = self.mutationSurveyorFormat(sampleInfo)
                    col_count += 1

        self.wb.save(self.newWbName)
        self.to_csv("MutationSurveyor", "txt", "\t")

    def mutationSurveyorFormat(self, sampleInfo):
        try:
            if sampleInfo.pNom != "" or sampleInfo.pNom is not None:
                amino = re.findall("[a-zA-Z]{3}", sampleInfo.pNom)
                resNo = re.findall("[0-9]+", sampleInfo.pNom)
                insNuc = "" if sampleInfo.insNucleotides is None else sampleInfo.insNucleotides
                print(amino[0], resNo[0])
        except:
            amino = ["?", "?"]
            resNo = ["?"]
            insNuc = "" if sampleInfo.insNucleotides is None else sampleInfo.insNucleotides

        if sampleInfo.cNom['cFull'] == "":
            return "Check variant"

        # Exonic substitution
        if sampleInfo.variantType == "substitution" and sampleInfo.variantLoc == "exon":
            if sampleInfo.genotype == "Heterozygous":
                return f"c.[{sampleInfo.cNom['cFull']}]+[=],p.{self.aa_conv(amino[0])}{resNo[0]}" \
                        + f"{self.aa_conv(amino[0])}{self.aa_conv(amino[1])}"
            elif sampleInfo.genotype == "Homozygous" or sampleInfo.genotype == "Hemizygous/Homozygous":
                return f"c.[{sampleInfo.cNom['cFull']}]+[{sampleInfo.cNom['cFull']}],p.{self.aa_conv(amino[0])}" \
                        + f"{resNo[0]}{self.aa_conv(amino[1])}"
        # Intronic substitution
        elif sampleInfo.variantType == "substitution" and sampleInfo.variantLoc.startswith(("intron", "splice")):
            if sampleInfo.genotype == "Heterozygous":
                return f"c.[{sampleInfo.cNom['cFull']}]+[=]"
            elif sampleInfo.genotype == "Homozygous" or sampleInfo.genotype == "Hemizygous/Homozygous":
                return f"c.[{sampleInfo.cNom['cFull']}]+[{sampleInfo.cNom['cFull']}]"
        # Frameshifts, insertions, deletions or duplications
        elif sampleInfo.codingEffect == "frameshift" or sampleInfo.variantType.startswith(("ins", "del", "dup")):
            if sampleInfo.genotype == "Heterozygous":
                if sampleInfo.cNom['cStart'] == sampleInfo.cNom['cEnd']:
                    return f"c.{sampleInfo.cNom['cFull'][:-3]}het_{sampleInfo.variantType[:3]}{insNuc}"
                else:
                    return f"c.{sampleInfo.cNom['cStart']}_{sampleInfo.cNom['cEnd']}het_{sampleInfo.variantType[:3]}{insNuc}"
            elif sampleInfo.genotype == "Homozygous" or sampleInfo.genotype == "Hemizygous/Homozygous":
                if sampleInfo.cNom['cStart'] == sampleInfo.cNom['cEnd']:
                    return f"c.{sampleInfo.cNom['cFull'][:-3]}het_{sampleInfo.variantType[:3]}{insNuc}"
                else:
                    return f"c.{sampleInfo.cNom['cStart']}_{sampleInfo.cNom['cEnd']}het_{sampleInfo.variantType[:3]}{insNuc}"
        else:
            return "Error"

    def to_csv(self, sheet, ext, delim):
        ws = self.wb.get_sheet_by_name(sheet)
        with open(f"{self.workDir}{self.varID}_{self.starID} {sheet} {self.timeNow}.{ext}", 'w', newline='') as f:
            c = csv.writer(f, delimiter=delim)
            for r in ws.rows:
                c.writerow([cell.value for cell in r])

    def closeWs(self):
        self.wb.close()






