import openpyxl
from datetime import datetime
import os
from npVariant import Variant

class Regex:

    def __init__(self, varPath, varID, workDir):
        self.varPath = varPath
        self.varID = varID
        self.workDir = workDir
        self.timeNow = datetime.now().strftime("%Y-%m-%d %H%M")
        self.newWbName = f"{self.workDir}{self.varID} regex {self.timeNow}.xlsx"
        self.wb = self.createCopy()
        self.ws_main = self.wb.get_sheet_by_name("STARLiMS_import")
        self.ws = self.createWorksheets()

        self.test()

    def createCopy(self):
        openpyxl.load_workbook(self.varPath).save(self.newWbName)
        return openpyxl.load_workbook(self.newWbName)

    def createWorksheets(self):
        # Create new sheet and populate header for mutation surveyor import file
        ws_ms = self.wb.create_sheet("MutationSurveyor")
        header_ms = ["Sample Name", "Reference Name", "Lane Quality", "ROI Coverage", "#nts below threshold",
                    "Quality ROI", "Variant1", "Variant2", "Variant3", "Variant4"]
        ws_ms.cell(row=1, column=1).value = "Warning!"
        for i in range(len(header_ms)):
            ws_ms.cell(row=2, column=i+1).value = header_ms[i]

        # Create new sheet and populate header for variant details import file
        ws_var = self.wb.create_sheet("VariantDetails")
        header_var = ["Well", "Sample", "VariantDetails",
                      "Gene1", "Zygosity1", "Inheritance1", "HGVS1", "Genomic1", "Pathogenicity1",
                      "Gene2", "Zygosity2", "Inheritance2", "HGVS2", "Genomic2", "Pathogenicity2"]
        for i in range(len(header_var)):
            ws_var.cell(row=1, column=i+1).value = header_var[i]
        self.wb.save(self.newWbName)
        return {'ms': ws_ms, 'var': ws_var}

    def test(self):
        var1 = Variant(self.ws_main, 21)
        var2 = Variant(self.ws_main, 31)
        var3 = Variant(self.ws_main, 32)
        var4 = Variant(self.ws_main, 34)





